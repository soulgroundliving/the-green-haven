#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract meter data from Excel files.
Processes three years of billing data (67, 68, 69)
Outputs JSON file ready for Firebase upload via Node.js script.
"""

import openpyxl
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

# Force UTF-8 output
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

# Configuration
EXCEL_FILES = {
    67: 'บิลปี67.xlsx',
    68: 'บิลปี68.xlsx',
    69: 'บิลปี69.xlsx',
}

# Note: Sheet 1 (EX) is reference/previous data, Sheets 2-13 = months 1-12

class MeterDataExtractor:
    def __init__(self):
        self.data = []
        self.errors = []

    def extract_from_excel(self, year: int, file_path: str):
        """Extract meter data from an Excel file"""
        print(f"Processing: {Path(file_path).name} (Year {year})")

        if not Path(file_path).exists():
            error_msg = f"File not found: {file_path}"
            print(error_msg)
            self.errors.append(error_msg)
            return

        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            print(f"  Total sheets: {len(wb.sheetnames)}")

            # Process sheets 2-13 (months 1-12)
            # Sheet 1 (index 0) is reference data (EX), skip it
            for sheet_idx in range(1, min(13, len(wb.sheetnames))):
                ws = wb.worksheets[sheet_idx]
                month = sheet_idx  # Sheet index 1-12 = Month 1-12

                # Extract data rows (skip header in row 1)
                rooms_processed = 0
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    if not row[0]:  # Skip empty rows
                        continue

                    room_id = self._normalize_room_id(row[0])
                    if not room_id:
                        continue

                    # Extract meter readings
                    # Column A (index 0) = Room ID
                    # Column B (index 1) = Water Current (New)
                    # Column C (index 2) = Water Previous (Old)
                    # Column L (index 11) = Electric Current (New)
                    # Column M (index 12) = Electric Previous (Old)

                    water_new = self._get_numeric_value(row[1])
                    water_old = self._get_numeric_value(row[2])

                    # Electric readings from columns L and M
                    electric_new = self._get_numeric_value(row[11]) if len(row) > 11 else None
                    electric_old = self._get_numeric_value(row[12]) if len(row) > 12 else None

                    # Skip if no valid water readings
                    if water_new is None or water_old is None:
                        continue

                    # Determine building
                    building = self._determine_building(room_id)

                    # Create document
                    doc_data = {
                        'building': building,
                        'year': year,
                        'month': month,
                        'roomId': str(room_id).strip(),
                        'wOld': water_old,
                        'wNew': water_new,
                        'eOld': electric_old if electric_old else 0,
                        'eNew': electric_new if electric_new else 0,  # FIXED: Use actual electric_new from column L, not water_old
                        'createdAt': datetime.now(timezone.utc).isoformat(),
                        'updatedAt': datetime.now(timezone.utc).isoformat(),
                    }

                    self.data.append(doc_data)
                    rooms_processed += 1

                print(f"    Month {month}: {rooms_processed} rooms")

            print(f"  OK: {Path(file_path).name}")

        except Exception as e:
            error_msg = f"Error processing {Path(file_path).name}: {str(e)}"
            print(error_msg)
            self.errors.append(error_msg)
            import traceback
            traceback.print_exc()

    def _normalize_room_id(self, room_id) -> str:
        """Normalize room ID format"""
        if room_id is None:
            return None

        room_str = str(room_id).strip()
        if not room_str:
            return None

        # Remove .0 from float strings like "13.0"
        if room_str.endswith('.0'):
            room_str = room_str[:-2]

        return room_str

    def _get_numeric_value(self, value):
        """Extract numeric value from cell (handles formulas by returning None)"""
        if value is None:
            return None

        # If it's a formula string, skip
        if isinstance(value, str) and value.startswith('='):
            return None

        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    def _determine_building(self, room_id: str) -> str:
        """Determine building type from room ID"""
        if not room_id:
            return 'rooms'

        first_char = room_id[0].upper()

        if first_char == 'N':
            return 'nest'
        else:
            return 'rooms'


    def save_to_json(self) -> bool:
        """Save extracted data to JSON file"""
        output_path = 'meter_data_export.json'

        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump({
                    'timestamp': datetime.now(timezone.utc).isoformat(),
                    'total_records': len(self.data),
                    'errors': self.errors,
                    'data': self.data
                }, f, indent=2, ensure_ascii=False)

            print(f"OK: Saved to {output_path}")
            return True
        except Exception as e:
            print(f"ERROR: Failed to save JSON: {str(e)}")
            return False

    def generate_report(self):
        """Generate summary report"""
        print("\n" + "="*60)
        print("EXTRACTION REPORT")
        print("="*60)

        print(f"\nTotal documents extracted: {len(self.data)}")

        # Group by year and month
        by_year_month = {}
        for doc in self.data:
            year = doc['year']
            month = doc['month']
            key = f"{year}_{month}"
            if key not in by_year_month:
                by_year_month[key] = 0
            by_year_month[key] += 1

        print("\nRecords by Year and Month:")
        for key in sorted(by_year_month.keys()):
            parts = key.split('_')
            print(f"  Year {parts[0]}, Month {parts[1]}: {by_year_month[key]} rooms")

        # Sample data
        if self.data:
            print("\nSample document (first entry):")
            sample = self.data[0]
            for key, value in sample.items():
                print(f"  {key}: {value}")

        if self.errors:
            print(f"\nErrors encountered: {len(self.errors)}")
            for error in self.errors[:5]:
                print(f"  - {error}")

        print("\n" + "="*60)


def main():
    """Main execution"""
    print("Meter Data Extraction")
    print("="*60)

    extractor = MeterDataExtractor()

    # Extract data from all Excel files
    for year, file_path in EXCEL_FILES.items():
        extractor.extract_from_excel(year, file_path)

    # Generate report and save
    extractor.generate_report()
    extractor.save_to_json()

    print("\nProcess complete!")


if __name__ == '__main__':
    main()
